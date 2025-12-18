import os, csv, io, re, zipfile, base64
from openpyxl import load_workbook
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
import textwrap
URL_REGEX = re.compile(r'(https?://\S+)')


# ---------------- READERS ---------------- #
from reportlab.lib.utils import simpleSplit

def draw_wrapped_text(c, text, x, y, max_width, font="Helvetica", size=10, leading=14):
    """
    Draw text with automatic wrapping & page-safe width
    Returns updated Y position
    """
    c.setFont(font, size)
    lines = simpleSplit(text, font, size, max_width)

    for line in lines:
        if y < 40:
            c.showPage()
            c.setFont(font, size)
            y = A4[1] - 40

        c.drawString(x, y, line)
        y -= leading

    return y


def read_csv(file):
    reader = csv.reader(io.TextIOWrapper(file, encoding="utf-8"))
    return [row for row in reader]


def read_excel(file):
    wb = load_workbook(file, read_only=True)
    sheet = wb.active
    return [[str(cell) if cell else "" for cell in row] for row in sheet.iter_rows(values_only=True)]


# ---------------- PDF BUILDERS ---------------- #

def draw_line(c, x, y, text):
    pos_x = x
    for part in re.split(f"({URL_REGEX.pattern})", text):
        c.setFillColor(colors.blue if re.match(URL_REGEX, part) else colors.black)
        c.drawString(pos_x, y, part)
        pos_x += c.stringWidth(part, "Helvetica", 10)


# -------- TYPE 1: TABULAR -------- #
def pdf_tabular(rows):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    x = 40
    y = height - 40
    max_width = width - 80

    # HEADER
    y = draw_wrapped_text(
        c,
        " | ".join(rows[0]),
        x, y,
        max_width,
        font="Helvetica-Bold",
        size=13
    )
    y -= 10

    # ROWS
    for row in rows[1:]:
        y = draw_wrapped_text(
            c,
            " | ".join(row),
            x, y,
            max_width
        )
        y -= 6

    c.save()
    buf.seek(0)
    return buf.read()



# -------- TYPE 2: RECORD / CARD -------- #
def pdf_card(rows):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    x = 40
    y = height - 40
    max_width = width - 80
    headers = rows[0]

    for row in rows[1:]:
        for i, value in enumerate(row):
            y = draw_wrapped_text(
                c,
                f"{headers[i]}: {value}",
                x, y,
                max_width
            )
        y -= 10
        c.line(x, y, width - 40, y)
        y -= 20

    c.save()
    buf.seek(0)
    return buf.read()



# -------- TYPE 3: COLUMNAR -------- #
def pdf_columnar(rows):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    x_margin = 40
    y_margin = 40
    y = height - y_margin
    line_height = 14
    max_width = width - (x_margin * 2)

    if not rows:
        c.drawString(x_margin, y, "No data found")
        c.save()
        buffer.seek(0)
        return buffer.getvalue()

    headers = rows[0]
    num_columns = len(headers)

    for col_idx in range(num_columns):

        # ---- HEADER ----
        c.setFont("Helvetica-Bold", 13)
        if y < y_margin:
            c.showPage()
            y = height - y_margin

        c.drawString(x_margin, y, f"{headers[col_idx]}")
        y -= line_height + 6

        # ---- COLUMN VALUES ----
        c.setFont("Helvetica", 10)

        for row in rows[1:]:
            value = row[col_idx] if col_idx < len(row) else ""
            value = str(value)

            wrapped_lines = textwrap.wrap(value, 90) or [""]

            for line in wrapped_lines:
                if y < y_margin:
                    c.showPage()
                    y = height - y_margin
                    c.setFont("Helvetica", 10)

                c.drawString(x_margin, y, line)
                y -= line_height

        y -= line_height  # spacing between columns

    c.save()
    buffer.seek(0)
    return buffer.getvalue()

# ---------------- AJAX VIEW ---------------- #

@csrf_exempt
def excel_to_pdf_ajax(request):
    print("\n===== START excel_to_pdf_ajax =====")

    files = request.FILES.getlist("files")
    layout = request.POST.get("layout")

    pdfs = []
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        for f in files:
            ext = os.path.splitext(f.name)[1].lower()
            rows = read_csv(f) if ext == ".csv" else read_excel(f)

            if layout == "tabular":
                pdf_bytes = pdf_tabular(rows)
            elif layout == "card":
                pdf_bytes = pdf_card(rows)
            elif layout == "columnar":
                pdf_bytes = pdf_columnar(rows)
            else:
                continue

            pdf_name = f"{os.path.splitext(f.name)[0]}-{layout}.pdf"
            zipf.writestr(pdf_name, pdf_bytes)

            pdfs.append({
                "file_name": pdf_name,
                "base64": "data:application/pdf;base64," +
                          base64.b64encode(pdf_bytes).decode()
            })

    zip_buffer.seek(0)

    print("===== DONE =====")

    return JsonResponse({
        "pdfs": pdfs,
        "zip_base64": "data:application/zip;base64," +
                      base64.b64encode(zip_buffer.read()).decode(),
        "zip_name": "Excel-to-PDFs.zip"
    })




import io
import csv
import zipfile
import base64
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook, load_workbook


@csrf_exempt
def convert_data(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)

    files = request.FILES.getlist("files")
    output_format = request.POST.get("outputFormat")

    print("===== START csv_excel_converter =====")
    print(f"Files: {len(files)} | Output: {output_format}")

    if not files or output_format not in ["csv", "xlsx"]:
        return JsonResponse({"error": "Invalid input"}, status=400)

    converted_files = []
    zip_buffer = io.BytesIO()
    zip_file = zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED)

    for idx, uploaded_file in enumerate(files, start=1):
        filename = uploaded_file.name
        ext = filename.split(".")[-1].lower()

        print(f"Processing file #{idx}: {filename}")

        try:
            # -------- READ INPUT --------
            if ext == "csv":
                text = uploaded_file.read().decode("utf-8")
                reader = csv.reader(io.StringIO(text))
                data = list(reader)

            elif ext in ["xls", "xlsx"]:
                wb = load_workbook(uploaded_file, read_only=True)
                sheet = wb.active
                data = [
                    [cell if cell is not None else "" for cell in row]
                    for row in sheet.iter_rows(values_only=True)
                ]
            else:
                continue

            # -------- WRITE OUTPUT --------
            output_buffer = io.BytesIO()

            if output_format == "csv":
                csv_buffer = io.StringIO()
                writer = csv.writer(csv_buffer)
                for row in data:
                    writer.writerow(row)

                output_buffer.write(csv_buffer.getvalue().encode("utf-8"))
                output_ext = "csv"
                mime = "text/csv"

            else:  # XLSX
                wb_out = Workbook()
                ws = wb_out.active
                for row in data:
                    ws.append(row)

                wb_out.save(output_buffer)
                output_ext = "xlsx"
                mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

            output_buffer.seek(0)

            output_name = filename.rsplit(".", 1)[0] + f".{output_ext}"

            # -------- BASE64 FOR AJAX --------
            encoded = base64.b64encode(output_buffer.read()).decode("utf-8")
            converted_files.append({
                "file_name": output_name,
                "base64": f"data:{mime};base64,{encoded}"
            })

            zip_file.writestr(output_name, base64.b64decode(encoded))

        except Exception as e:
            print("ERROR:", e)
            return JsonResponse({"error": "Conversion failed"}, status=400)

    zip_file.close()
    zip_buffer.seek(0)

    zip_base64 = base64.b64encode(zip_buffer.read()).decode("utf-8")

    print("===== DONE =====")

    return JsonResponse({
        "pdfs": converted_files,
        "zip_name": "converted_files.zip",
        "zip_base64": f"data:application/zip;base64,{zip_base64}"
    })
